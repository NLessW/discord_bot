import discord
import datetime 
import openpyxl
import requests
import location
import asyncio
import random
import urllib
import bs4
import youtube_dl
import json
import os
from discord.ext import commands
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
from datetime import datetime


############봇 상태############
@client.event 
async def on_ready():
    print(client.user.id)
    print("봇이 켜져버려욧!")
    game = discord.Game("봇 도움말 !nh help")
    await client.change_presence(status=discord.Status.online, activity=game)

############인사 말############
@client.event
async def on_message(message): 
    await client.process_commands(message)

############도움말############
    if message.content.startswith("!nh help") and not message.content.startswith("!nh help 1") and not message.content.startswith("!nh help 2"):
        embed = discord.Embed(color = 0x900020, title = "!nh help (페이지 1, 2선택)" , desciption = "도움말이 많아 2페이지로 나누었습니다.")
        await message.channel.send(embed=embed)

    if message.content.startswith("!nh help 1"):
        embed = discord.Embed(color=0x900020, title = '1 페이지')
        embed.add_field(name="업데이트", value = "명령어 : !nh업데이트\n최신 업데이트 1개를 보여줍니다.", inline = False)
        embed.add_field(name="인사", value = "명령어 : !nh안녕\n인사를 해줍니다.", inline = False)
        embed.add_field(name="학습", value = "명령어 : !nh학습 (학습시킬 단어) (출력단어) 출력단어에는 띄어쓰기하면 안됩니다.")
        embed.add_field(name="기억", value = "명령어 : !nh기억 (학습시킨 단어)", inline= False)
        embed.add_field(name="봇 핑 확인", value = "명령어 : !nh ping\n봇의 전송속도 핑을 보여줍니다.", inline = False)
        embed.add_field(name="프로필", value = "명령어 : !nh정보\n본인의 디스코드 프로필을 보여줍니다.", inline = False)
        embed.add_field(name="코로나", value = "명령어 : !nh코로나\n현재 대한민국의 코로나 현황을 보여줍니다.", inline = False)
        embed.add_field(name="날씨", value = "명령어 : !nh날씨 (지역)\n검색한 지역의 날씨를 보여줍니다.", inline = False)
        await message.channel.send(embed=embed)

    if message.content.startswith("!nh help 2"):
        embed = discord.Embed(color=0x900020, title = '2 페이지')
        embed.add_field(name="팀나누기", value = "명령어 : !nh팀나누기 (나눌사람의 이름[공백으로 구분]/(팀이름))\n추첨기나 팀 나눌때 사용합니다\n(ex/a b c d를 1팀과 2팀으로 나눌경우 !nh팀나누기 a b c d/1 2 1 2)팀은 꼭 인원수 만큼 써주세요!", inline = False)
        embed.add_field(name="한강수온", value = "명령어 : !nh한강\n한강의 현재 수온을 보여줍니다.", inline = False)
        embed.add_field(name="주사위", value = "명령어 : !nh주사위 돌릴횟수d면갯수\n주사위를 n번만큼 굴려 합을 구해줍니다.\nex)!nh주사위 3d6 = 6면체주사위를 3번 굴린다.", inline = False)
        embed.add_field(name="카트라이더 전적", value = "명령어 : !nh카트 (닉네임)\n검색한 유저의 전적을 보여줍니다.", inline = False)
        embed.add_field(name="롤 솔로랭크 전적", value =  "명령어 : !nh롤솔랭 (닉네임)\n언랭은 검색해도 나오지 않습니다.\n만약 나오지 않을 시 poro.gg나 op.gg가셔서 전적갱신을 해주시기 바랍니다.", inline=False)
        embed.add_field(name="롤 언랭 전적", value = "명령어 : !nh롤언랭 (닉네임)\n언랭들을 위한 전적검색입니다.\n만약 나오지 않을 시 poro.gg나 op.gg가셔서 전적갱신을 해주시기 바랍니다.", inline=False)
        embed.add_field(name="노래", value = "명령어 : !nh재생\n보이스 채널에 연결합니다.\n아직 재생은 구현중입니다.", inline = False)
        embed.add_field(name="연결끊기", value = "명령어 : !nh연결끊기\n보이스 채널에 있는 봇의 연결을 끊습니다.", inline = False)
        await message.channel.send(embed=embed)

    if message.content.startswith("!nh안녕"):
        await message.channel.send("ㅎㅇ")

############핑 확인############
    if message.content.startswith("!nh ping"):
        latancy = client.latency
        await message.channel.send(f'Ping : {round(latancy * 1000)}ms') 

############디스코드 프로필 확인############
    if message.content.startswith("!nh정보"):
        date = datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000)/1000)
        embed = discord.Embed(color=0x900020)
        embed.add_field(name="이름", value=message.author.name, inline=False)
        embed.add_field(name="서버 닉네임", value=message.author.display_name, inline=False)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월 " + str(date.day) + "일", inline=False)
        embed.add_field(name="ID", value=message.author.id, inline=False)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)

############말 가르치기############
    if message.content.startswith("!nh학습"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]: 
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("단어가 학습되었습니다.")
                break
        file.save("기억.xlsx")
    
    if message.content.startswith("!nh기억") and not message.content.startswith("!nh기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break
    
    if message.content.startswith("!nh기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await message.channel.send("기억이 삭제되었습니다.")
                file.save("기억.xlsx")
                break
    
   

############사전############
    if message.content.startswith("!nh사전 "):
        response = requests.get('https://dict.naver.com/search.nhn?dicQuery='+ message.content[5:])
        readerhtml = response.text
        soup = BeautifulSoup(readerhtml, 'lxml')
        cla = soup.find('div', {'class' : 'sta'})
        ulfind = cla.find('ul', {'class' : 'lst_krdic'})
        lifind = ulfind.findAll('br')
        mean = lifind[1].text
        embed=discord.Embed(color = 0x900020, title = message.content[5:])
        embed.add_field(name = "뜻", value = mean)
        await message.channel.send(embed=embed)


############날씨############
    if message.content.startswith("!nh날씨 "):
        response = requests.get('https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=' + message.content[5:] +'날씨')
        readerhtml = response.text
        soup = BeautifulSoup(readerhtml, 'lxml')

        NowTemp = soup.find('span', {'class': 'todaytemp'}).text + soup.find('span', {'class' : 'tempmark'}).text[2:]        
        WeatherCast = soup.find('p', {'class' : 'cast_txt'}).text
        TodayMorningTemp = soup.find('span', {'class' : 'min'}).text
        TodayAfternoonTemp = soup.find('span', {'class' : 'max'}).text
        TodayFeelTemp = soup.find('span', {'class' : 'sensible'}).text[5:]
        TodayUV = soup.find('span', {'class' : 'indicator'}).text[4:-2] + " " + soup.find('span', {'class' : 'indicator'}).text[-2:]
        data1 = soup.find('div',{'class':'detail_box'})
        data2 = data1.findAll('dd')
        FineDust = data2[0].find('span',{'class':'num'}).text
        UltraFineDust = data2[1].find('span',{'class' : 'num'}).text
        Ozon = data2[2].find('span', {'class':'num'}).text

        embed = discord.Embed(color=0x900020, title = '🌞'+message.content[5:]+'의 날씨🌞')
        embed.add_field(name="현재 날씨", value=NowTemp, inline = False)
        embed.add_field(name="현재 상태 / 어제와 오늘", value=WeatherCast, inline = False)
        embed.add_field(name="최저/최고", value=f'{TodayMorningTemp}C' + "/" f'{TodayAfternoonTemp}C', inline = False)
        embed.add_field(name="오늘 체감 온도", value=f'{TodayFeelTemp}C',inline = False)
        embed.add_field(name="자외선", value=TodayUV, inline = False)
        embed.add_field(name="미세먼지", value=FineDust)
        embed.add_field(name="초미세먼지", value=UltraFineDust)
        embed.add_field(name="오존", value=Ozon)
        embed.set_footer(text="Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed=embed)
############팀나누기############
    if message.content.startswith("!nh팀나누기"):
        team = message.content[8:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)

        embed = discord.Embed(color=0x900020, title = "팀 나누기 결과")
        for i in range(0, len(person)):
            embed.add_field(name="결과", value=person[i] + " ----> " + teamname[i], inline = False)
        embed.set_footer(text = "Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed=embed)

############주사위############
    if message.content.startswith("!nh주사위"):
        roll = message.content.split(" ")
        rolld = roll[1].split("d")
        dice = 0
        for i in range(1, int(rolld[0])+1):
            dice = dice + random.randint(1, int(rolld[1]))
        #await message.channel.send(str(dice))
        embed = discord.Embed(color=0x900020, title = "🎲주사위 굴리기🎲")
        embed.add_field(name=roll[1][2:]+"면체 주사위를 "+roll[1][:1]+"번 굴려 나온 주사위 합", value = str(dice))
        embed.set_footer(text = "Source - EstelBlHERO")
        await message.channel.send(embed=embed)
    

############카트############
    if message.content.startswith("!nh카트"):
        response = requests.get('http://kart.nexon.com/Garage/Main?strRiderID='+message.content[6:])
        response2 = requests.get('http://kart.nexon.com/Garage/Record?strRiderID='+message.content[6:])
        
        #크롤링 파일 형식#
        readerhtml = response.text
        readerhtml2 = response2.text
        
        #크롤링#
        soup = BeautifulSoup(readerhtml, 'lxml')
        soup2 = BeautifulSoup(readerhtml2, 'lxml')
         
        #차고1#
        nick = soup.find('span', {'id' : 'RiderName'}).text
        club = soup.find('span', {'id' : 'GuildName'}).text
        rprank = soup.find('span',{'class' : 'RecordData1'}).text
        rp = soup.find('span',{'class' : 'RecordData2'}).text
        avatar = soup.find('div', {'id' : 'CharInfo'})
        avatar2 = avatar.find('img').get('src')
        glove = soup.find('div', {'id' : 'GloveImg'})
        glove2 = glove.find('img').get('src')
        #차고2#
        cnt = soup2.find('div', {'id' : 'CntRecord2'})
        dlfind = cnt.findAll('dl')
        starty = dlfind[0].find('dd').text[0:4]
        startm = dlfind[0].find('dd').text[5:7]
        startd = dlfind[0].find('dd').text[8:10]
        startday = dlfind[0].find('dd').text[11:]
        racing = dlfind[1].find('dd').text
        gameon = dlfind[2].find('dd').text
        #최근 접속#
        recenty = dlfind[3].find('dd').text[0:4]
        recentm = dlfind[3].find('dd').text[5:7]
        recentd = dlfind[3].find('dd').text[8:10]        
        #전체 승률#
        recorddata2 = soup2.find('div', {'id' : 'CntRecord'})
        allwinrate = recorddata2.find('td',{'class' : 'RecordL2'}).text[0:3]
        allwin = recorddata2.find('td',{'class' : 'RecordL2'}).text[4:]
        allwinrp = recorddata2.find('td',{'class' : 'RecordL3'}).text       
        #스피드#
        winrate = recorddata2.find('table', {'class' : 'RecordL'})
        sprate = winrate.findAll('td')
        spallrt = sprate[4].text[0:3]
        spallrt2 = sprate[4].text[4:]
        sprprank = sprate[5].text
        #아이템#
        iprallrt = sprate[7].text[0:3]
        iprallrt2 = sprate[7].text[4:]
        iprprank = sprate[8].text
        #출력#
        kartembed = discord.Embed(color=0x900020)
        kartembed.set_author(name= nick, icon_url= glove2)
        kartembed.add_field(name = "Club", value = club, inline = True)
        kartembed.add_field(name = "RP", value = rprank + "\n" + rp, inline = True)
        kartembed.add_field(name = "All Win Rate", value = allwinrate + "\n" + "(" + allwin + ")", inline = True)
        kartembed.add_field(name = "Speed Win Rate", value = spallrt + "\n" + "(" + spallrt2 + ")", inline = True)
        kartembed.add_field(name = "Item Win Rate", value = iprallrt + "\n" + "(" + iprallrt2 + ")", inline = True)
        kartembed.add_field(name = "All RP", value = allwinrp, inline = True)
        kartembed.add_field(name = "Speed RP", value = sprprank, inline = True)
        kartembed.add_field(name = "Item RP", value = iprprank, inline = True)
        kartembed.add_field(name = "Rider Creation", value = f'{starty}년 '+f'{startm}월 '+f'{startd}일' "\n" + startday, inline = True)
        kartembed.add_field(name = "Driving Time", value = racing, inline = True)
        kartembed.add_field(name = "Game Runs", value = gameon, inline = True)
        kartembed.add_field(name = "Recent Access", value = f'{recenty}년 '+f'{recentm}월 '+f'{recentd}일')
        kartembed.add_field(name="TMI",value=f'[KartRiderTMI](https://tmi.nexon.com/kart/user?nick={nick})')
        kartembed.set_thumbnail(url = avatar2)
        await message.channel.send(embed=kartembed)


############롤 솔랭############        
    if message.content.startswith("!nh롤솔랭"):
        response = requests.get('https://www.op.gg/summoner/userName='+message.content[7:]) #롤 전적사이트 op.gg 링크
        response2 = requests.get('https://poro.gg/ko/s/KR/'+message.content[7:])
        readerhtml = response.text
        readerhtml2 = response2.text

        soup = BeautifulSoup(readerhtml, 'lxml')
        soup2 = BeautifulSoup(readerhtml2, 'lxml')

        profile = soup.find('div',{'class' : 'Profile'}) #profile 크롤링
        nickname = profile.findAll('span') #닉네임
        nick = nickname[1].text #닉네임
        leve = soup.find('div', {'class' : 'Face'})#레벨
        lev = leve.findAll('span')
        level = lev[0].text
        
        #솔랭
        win = soup.find('span',{'class' : 'wins'}).text[0:3]
        lose = soup.find('span',{'class' : 'losses'}).text[0:3]
        winrate1 = soup2.find('div', {'class' : 'summoner__tier__winrate text-gray'})
        winrate = winrate1.find('b').text
        medal = soup2.find('div', {'class' : 'rank-info__tier'})
        medalimg = medal.find('img').get('src')

        tier1 = soup2.find('div',{'class' : 'summoner__tier__content'})
        tier = tier1.find('b').text
        
        recentrank = soup2.find('div', {'class' : 'recent-match-condition__summary'})
        recentrate = recentrank.find('span', {'class' : 'recent-match-condition__summary__winrate'}).text
        recentwinlose = recentrank.find('span',{'class' : 'recent-match-condition__summary__winrate-text'}).text
        update = soup2.find('div', {'class' : 'summoner-header__profile__updated'}).text

        lolembed = discord.Embed(color=0x900020)
        lolembed.set_author(name = message.content[7:]+"님의 솔로랭크입니다.")
        lolembed.add_field(name = "레벨", value = f'{level}레벨', inline = True)            
        lolembed.add_field(name = "전적", value = f'{win}승 '+"/ "+ f'{lose}패', inline = True)
        lolembed.add_field(name = "승률", value = f'{winrate}%', inline = True)            
        lolembed.add_field(name = "티어", value = tier, inline = True)
        lolembed.add_field(name = "최근 랭크게임", value = recentrate+"\n"+recentwinlose, inline = True)            
        lolembed.set_thumbnail(url = medalimg)
        lolembed.set_footer(text = update + "\nSource - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed = lolembed)
    
############롤 언랭############
    if message.content.startswith("!nh롤언랭"):
        response = requests.get('https://poro.gg/ko/s/KR/'+message.content[7:])
        response2 = requests.get('https://www.op.gg/summoner/userName='+message.content[7:])

        readerhtml = response.text
        readerhtml2 = response2.text

        soup = BeautifulSoup(readerhtml, 'lxml')
        soup2 = BeautifulSoup(readerhtml2, 'lxml')

        nick1 = soup.find('div', {'class' : 'summoner-header__profile__info'})
        nick = nick1.find('h3').text
        level = soup.find('div', {'class' : 'summoner-header__profile__level'}).text

        winratio = soup2.find('div',{'class' : 'WinRatioTitle'})
        spanfind = winratio.findAll('span')
        win = spanfind[1].text
        lose = spanfind[2].text
        update = soup.find('div', {'class' : 'summoner-header__profile__updated'}).text
        
        medal = soup.find('div', {'class' : 'rank-info__tier'})
        medalimg = medal.find('img').get('src')
        
        lolembed = discord.Embed(color=0x900020)
        lolembed.set_author(name = "언랭"+nick+"님 정보입니다.")
        lolembed.add_field(name = "레벨", value = f'{level}레벨', inline = True)
        lolembed.add_field(name = "최근 20게임 전적", value = f'{win}승 ' +"/ "+ f'{lose}패', inline = True)
        lolembed.add_field(name = "랭크를 돌립시다!", value = "랭크가 생기면 더 많은 정보를 받아올 수 있어요!", inline = False)
        lolembed.set_footer(text=update + "\nSource - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        lolembed.set_thumbnail(url = medalimg)
        await message.channel.send(embed = lolembed)





    if message.content.startswith("!nh공지"):
        file = openpyxl.load_workbook("서버목록.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                channel = sheet["B" + str(i)].value
                msg = message.content[6:]
                foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
                embed = discord.Embed(title="넥히봇 공지", description=msg, color=0x900020)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                embed.set_thumbnail(url='https://cdn.discordapp.com/avatars/688686602804920364/7ca067359e2235dd6c817480adef9075.png?size=128')
                embed.set_footer(text=foot + "봇 관련 문의는 Peto#6092")  # 하단에 들어가는 조그마한 설명을 잡아줍니다
                await client.get_channel(int(channel)).send(embed=embed)
                break
            if sheet["A" + str(i)].value == None:
                await message.channel.send("공지채널이 설정되어 있지 않습니다.")
                break
            i += 1

    if message.content.startswith("!nh설정"):
        channel = int(message.content[6:26])
        file = openpyxl.load_workbook("서버목록.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                sheet["B" + str(i)].value = str(channel)
                file.save("서버목록.xlsx")
                await message.channel.send("정상적으로 설정되었습니다!")
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.guild.id)
                sheet["B" + str(i)].value = str(channel)
                file.save("서버목록.xlsx")
                await message.channel.send("정상적으로 설정되었습니다!")
                break
            i += 1

    



    if message.content.startswith("!nh업데이트"):
        embed = discord.Embed(color=0x900020, title = "업데이트")
        embed.add_field(name = "2020년 4월 13일 업데이트 내용입니다.", value = "!nh help명령어 사용법이 변경되었습니다. !nh help (페이지 쪽 수)로 업데이트 되었습니다.\n롤 전적 검색 기능이 추가되었습니다. !nh롤솔랭 (닉네임), !nh롤언랭 (닉네임)으로 검색가능합니다.")
        embed.set_footer(text="봇과 관련된 문의는 Peto#6092")

        await message.channel.send(embed=embed)


client.run(os.environ["token"])
