import discord
import datetime 
import openpyxl
import requests
import request
import location
import asyncio
import random
import urllib
import bs4
import youtube_dl
import json
import os
from discord.ext import commands
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
from datetime import datetime

client = commands.Bot(command_prefix='!nh') 

############봇 상태############
@client.event 
async def on_ready():
    print(client.user.id)
    print("봇이 켜져버려욧!")
    game = discord.Game("봇 도움말 !nh help")
    await client.change_presence(status=discord.Status.online, activity=game)

############인사 말############
@client.event
async def on_message(message): 
    await client.process_commands(message)

############도움말############
    if message.content.startswith("!nh help"):
        embed = discord.Embed(color=0x900020, title = '도움말')
        embed.add_field(name="업데이트", value = "명령어 : !nh업데이트\n최신 업데이트 1개를 보여줍니다.", inline = False)
        embed.add_field(name="인사", value = "명령어 : !nh안녕\n인사를 해줍니다.", inline = False)
        embed.add_field(name="봇 핑 확인", value = "명령어 : !nh ping\n봇의 전송속도 핑을 보여줍니다.", inline = False)
        embed.add_field(name="프로필", value = "명령어 : !nh정보\n본인의 디스코드 프로필을 보여줍니다.", inline = False)
        embed.add_field(name="코로나", value = "명령어 : !nh코로나\n현재 대한민국의 코로나 현황을 보여줍니다.", inline = False)
        embed.add_field(name="날씨", value = "명령어 : !nh날씨 (지역)\n검색한 지역의 날씨를 보여줍니다.", inline = False)
        embed.add_field(name="팀나누기", value = "명령어 : !nh팀나누기 (나눌사람의 이름[공백으로 구분]/(팀이름))\n추첨기나 팀 나눌때 사용합니다\n(ex/a b c d를 1팀과 2팀으로 나눌경우 !nh팀나누기 a b c d/1 2 1 2)팀은 꼭 인원수 만큼 써주세요!", inline = False)
        embed.add_field(name="한강수온", value = "명령어 : !nh한강\n한강의 현재 수온을 보여줍니다.", inline = False)
        embed.add_field(name="주사위", value = "명령어 : !nh주사위 돌릴횟수d면갯수\n주사위를 n번만큼 굴려 합을 구해줍니다.\nex)!nh주사위 3d6 = 6면체주사위를 3번 굴린다.", inline = False)
        embed.add_field(name="카트라이더 전적", value = "명령어 : !nh카트 (닉네임)\n검색한 유저의 전적을 보여줍니다.", inline = False)
        embed.add_field(name="노래", value = "명령어 : !nh재생\n보이스 채널에 연결합니다.\n아직 재생은 구현중입니다.", inline = False)
        embed.add_field(name="연결끊기", value = "명령어 : !nh연결끊기\n보이스 채널에 있는 봇의 연결을 끊습니다.", inline = False)
        await message.channel.send(embed=embed)

    if message.content.startswith("!nh안녕"):
        await message.channel.send("ㅎㅇ")

############핑 확인############
    if message.content.startswith("!nh ping"):
        latancy = client.latency
        await message.channel.send(f'Ping : {round(latancy * 1000)}ms') 

############디스코드 프로필 확인############
    if message.content.startswith("!nh정보"):
        date = datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000)/1000)
        embed = discord.Embed(color=0x900020)
        embed.add_field(name="이름", value=message.author.name, inline=False)
        embed.add_field(name="서버 닉네임", value=message.author.display_name, inline=False)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월 " + str(date.day) + "일", inline=False)
        embed.add_field(name="ID", value=message.author.id, inline=False)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)

############말 가르치기############
    if message.content.startswith("!nh학습"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]: 
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("단어가 학습되었습니다.")
                break
        file.save("기억.xlsx")
    
    if message.content.startswith("!nh기억") and not message.content.startswith("!nh기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break
    
    if message.content.startswith("!nh기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await message.channel.send("기억이 삭제되었습니다.")
                file.save("기억.xlsx")
                break
    
   


############코로나############
    if message.content.startswith("!nh코로나"):
        response = requests.get('https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=코로나')
        readerhtml = response.text
        soup = BeautifulSoup(readerhtml, 'lxml')
        data1 = soup.find('div', class_='graph_view')
        data2 = data1.findAll('div', class_='box')
        data3 = data1.findAll('div', class_='box bottom')
        checked = data2[0].find('p', class_='txt').find('strong', class_='num').text
        checking = data2[2].find('p', class_='txt').find('strong', class_='num').text
        free = data3[0].find('p', class_='txt').find('strong', class_='num').text        
        die = data3[1].find('p', class_='txt').find('strong', class_='num').text
        wasup = soup.find('div', class_='csp_notice_info').find('p').find_all(text=True, recursive=True)
        
        coembed = discord.Embed(color=0x900020, title='☣코로나현황☣', description =f'{wasup[1]}' )
        coembed.add_field(name="☣확진자☣", value=f'{checked}명', inline=False)
        coembed.add_field(name="💉격리해제💉", value=f'{free}명', inline=False)
        coembed.add_field(name="🔎검사중🔎", value=f'{checking}명', inline=False)
        coembed.add_field(name="👻사망자👻", value=f'{die}명', inline=False)                
        coembed.set_footer(text="Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed = coembed)

############날씨############
    if message.content.startswith("!nh날씨 "):
        response = requests.get('https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=' + message.content[5:] +'날씨')
        readerhtml = response.text
        soup = BeautifulSoup(readerhtml, 'lxml')

        NowTemp = soup.find('span', {'class': 'todaytemp'}).text + soup.find('span', {'class' : 'tempmark'}).text[2:]        
        WeatherCast = soup.find('p', {'class' : 'cast_txt'}).text
        TodayMorningTemp = soup.find('span', {'class' : 'min'}).text
        TodayAfternoonTemp = soup.find('span', {'class' : 'max'}).text
        TodayFeelTemp = soup.find('span', {'class' : 'sensible'}).text[5:]
        TodayUV = soup.find('span', {'class' : 'indicator'}).text[4:-2] + " " + soup.find('span', {'class' : 'indicator'}).text[-2:]
        data1 = soup.find('div',{'class':'detail_box'})
        data2 = data1.findAll('dd')
        FineDust = data2[0].find('span',{'class':'num'}).text
        UltraFineDust = data2[1].find('span',{'class' : 'num'}).text
        Ozon = data2[2].find('span', {'class':'num'}).text

        embed = discord.Embed(color=0x900020, title = '🌞'+message.content[5:]+'의 날씨🌞')
        embed.add_field(name="현재 날씨", value=NowTemp, inline = False)
        embed.add_field(name="현재 상태 / 어제와 오늘", value=WeatherCast, inline = False)
        embed.add_field(name="최저/최고", value=f'{TodayMorningTemp}C' + "/" f'{TodayAfternoonTemp}C', inline = False)
        embed.add_field(name="오늘 체감 온도", value=f'{TodayFeelTemp}C',inline = False)
        embed.add_field(name="자외선", value=TodayUV, inline = False)
        embed.add_field(name="미세먼지", value=FineDust)
        embed.add_field(name="초미세먼지", value=UltraFineDust)
        embed.add_field(name="오존", value=Ozon)
        embed.set_footer(text="Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed=embed)
############팀나누기############
    if message.content.startswith("!nh팀나누기"):
        team = message.content[8:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)

        embed = discord.Embed(color=0x900020, title = "팀 나누기 결과")
        for i in range(0, len(person)):
            embed.add_field(name="결과", value=person[i] + " ----> " + teamname[i], inline = False)
        embed.set_footer(text = "Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed=embed)

############한강############
    if message.content.startswith("!nh한강"):
        response = requests.get('https://www.wpws.kr/hangang/')
        readerhtml = response.text
        soup = BeautifulSoup(readerhtml, 'lxml')
        tempdata1 = soup.find('p', {'id' : 'temp'}).text
        tempdata2 = soup.find('p', {'id' : 'foo2'}).text[2:21]
        embed = discord.Embed(color=0x900020, title = "💧현재 한강의 온도💧")
        embed.add_field(name="🌡온도", value=tempdata1, inline = False)
        embed.add_field(name="⌛측정 시간", value=tempdata2, inline = False)
        embed.set_footer(text="📞자살예방상담전화 : 1393\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed=embed)

############주사위############
    if message.content.startswith("!nh주사위"):
        roll = message.content.split(" ")
        rolld = roll[1].split("d")
        dice = 0
        for i in range(1, int(rolld[0])+1):
            dice = dice + random.randint(1, int(rolld[1]))
        #await message.channel.send(str(dice))
        embed = discord.Embed(color=0x900020, title = "🎲주사위 굴리기🎲")
        embed.add_field(name=roll[1][2:]+"면체 주사위를 "+roll[1][:1]+"번 굴려 나온 주사위 합", value = str(dice))
        embed.set_footer(text = "Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        await message.channel.send(embed=embed)
    

############카트############
    if message.content.startswith("!nh카트"):
        response = requests.get('http://kart.nexon.com/Garage/Main?strRiderID='+message.content[5:])
        response2 = requests.get('http://kart.nexon.com/Garage/Record?strRiderID='+message.content[5:])
       
        readerhtml = response.text
        readerhtml2 = response2.text

        soup = BeautifulSoup(readerhtml, 'lxml')
        soup2 = BeautifulSoup(readerhtml2, 'lxml')

        #차고1#
        nick = soup.find('span', {'id' : 'RiderName'}).text
        club = soup.find('span', {'id' : 'GuildName'}).text
        rprank = soup.find('span',{'class' : 'RecordData1'}).text
        rp = soup.find('span',{'class' : 'RecordData2'}).text
        avatar = soup.find('div', {'id' : 'CharInfo'})
        avatar2 = avatar.find('img').get('src')
        
        #차고2#
        cnt = soup2.find('div', {'id' : 'CntRecord2'})
        dlfind = cnt.findAll('dl')
        starty = dlfind[0].find('dd').text[0:4]
        startm = dlfind[0].find('dd').text[5:7]
        startd = dlfind[0].find('dd').text[8:10]
        startday = dlfind[0].find('dd').text[11:]
        racing = dlfind[1].find('dd').text
        gameon = dlfind[2].find('dd').text

        recenty = dlfind[3].find('dd').text[0:4]
        recentm = dlfind[3].find('dd').text[5:7]
        recentd = dlfind[3].find('dd').text[8:10]

        
        #전체 승률#
        recorddata2 = soup2.find('div', {'id' : 'CntRecord'})
        allwinrate = recorddata2.find('td',{'class' : 'RecordL2'}).text[0:3]
        allwin = recorddata2.find('td',{'class' : 'RecordL2'}).text[4:]
        allwinrp = recorddata2.find('td',{'class' : 'RecordL3'}).text
        
        #스피드#
        winrate = recorddata2.find('table', {'class' : 'RecordL'})
        sprate = winrate.findAll('td')
        spallrt = sprate[4].text[0:3]
        spallrt2 = sprate[4].text[4:]
        sprprank = sprate[5].text
        
        #아이템#
        iprallrt = sprate[7].text[0:3]
        iprallrt2 = sprate[7].text[4:]
        iprprank = sprate[8].text

        embed = discord.Embed(color=0x900020, title = message.content[5:])
        embed.add_field(name = "NickName", value = nick, inline = True)
        embed.add_field(name = "Club", value = club, inline = True)
        embed.add_field(name = "RP", value = rprank + "\n" + rp, inline = True)
        embed.add_field(name = "All Win Rate", value = allwinrate + "\n" + "(" + allwin + ")", inline = True)
        embed.add_field(name = "Speed Win Rate", value = spallrt + "\n" + "(" + spallrt2 + ")", inline = True)
        embed.add_field(name = "Item Win Rate", value = iprallrt + "\n" + "(" + iprallrt2 + ")", inline = True)
        embed.add_field(name = "All RP", value = allwinrp, inline = True)
        embed.add_field(name = "Speed RP", value = sprprank, inline = True)
        embed.add_field(name = "Item RP", value = iprprank, inline = True)
        embed.add_field(name = "Rider Creation", value = f'{starty}년 '+f'{startm}월 '+f'{startd}일' "\n" + startday, inline = True)
        embed.add_field(name = "Driving Time", value = racing, inline = True)
        embed.add_field(name = "Game Runs", value = gameon, inline = True)
        embed.add_field(name = "Recent Access", value = f'{recenty}년 '+f'{recentm}월 '+f'{recentd}일')
        embed.add_field(name="TMI",value=f'[KartRiderTMI](https://tmi.nexon.com/kart/user?nick={nick})')
        embed.set_footer(text="Source - NextHeroes\nLv2 S2 KartRiderClub NextLv's Bot")
        embed.set_thumbnail(url = avatar2)
        await message.channel.send(embed=embed)
        
            

    if message.content.startswith("!nh공지"):
        file = openpyxl.load_workbook("서버목록.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                channel = sheet["B" + str(i)].value
                msg = message.content[6:]
                foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
                embed = discord.Embed(title="넥히봇 공지", description=msg, color=0x900020)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                embed.set_thumbnail(url='https://cdn.discordapp.com/avatars/688686602804920364/7ca067359e2235dd6c817480adef9075.png?size=128')
                embed.set_footer(text=foot + "봇 관련 문의는 Peto#6092")  # 하단에 들어가는 조그마한 설명을 잡아줍니다
                await client.get_channel(int(channel)).send(embed=embed)
                break
            if sheet["A" + str(i)].value == None:
                await message.channel.send("공지채널이 설정되어 있지 않습니다.")
                break
            i += 1

    if message.content.startswith("!nh설정"):
        channel = int(message.content[6:26])
        file = openpyxl.load_workbook("서버목록.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                sheet["B" + str(i)].value = str(channel)
                file.save("서버목록.xlsx")
                await message.channel.send("정상적으로 설정되었습니다!")
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.guild.id)
                sheet["B" + str(i)].value = str(channel)
                file.save("서버목록.xlsx")
                await message.channel.send("정상적으로 설정되었습니다!")
                break
            i += 1

    if message.content.startswith("!nh업데이트"):
        embed = discord.Embed(color=0x900020, title = "업데이트")
        embed.add_field(name = "2020년 4월 9일 업데이트 내용입니다.", value = "카트 전적봇의 TMI 링크가 추가되었습니다")
        embed.set_footer(text="봇과 관련된 문의는 Peto#6092")

        await message.channel.send(embed=embed)


    file = openpyxl.load_workbook("레벨.xlsx")
    sheet = file.active
    exp = [10, 20, 30, 40, 50]
    i = 1
    while True:
        if sheet["A" + str(i)].value == str(message.author.id):
            sheet["B" + str(i)].value == sheet["B" + str(i)].value + 5
            if sheet["B" + str(i)].value >= exp[sheet["C" + str(i)].value]:
                sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                await message.channel.send("레벨이 올랐습니다.\n현재 레벨 : " + str(sheet["C" + str(i)].value) + "\n경험치 : " + str(sheet["B" + str(i)].value))
            file.save("레벨.xlsx")
            break
            
            
            
        if sheet["A" + str(i)].value == None:
            sheet["A" + str(i)].value = str(message.author.id)
            sheet["B" + str(i)].value = 0
            sheet["C" + str(i)].value = 1
            file.save("레벨.xlsx")
            break

        i += 1 

    
    
@client.command(name="재생", pass_context=True)
async def _join(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
        await channel.connect()
    else:
        await ctx.send("채널에 연결되지 않았습니다.")

    
@client.command(name="연결끊기")
async def _leave(ctx):
    await client.voice_clients[0].disconnect()

client.run("Token")
